import openpyxl
import random
from datetime import datetime, timedelta

# Define the possible values for each field
statuses = ['New', 'In-progress', 'Resolved', 'Closed']
priorities = [1, 2, 3, 4, 5]
weights = [1, 2, 3, 4, 5]

# Create a new workbook and select the active worksheet
workbook = openpyxl.Workbook()
worksheet = workbook.active

# Define the headers for each column
worksheet['A1'] = 'Incident Number'
worksheet['B1'] = 'Priority'
worksheet['C1'] = 'Date/Time'
worksheet['D1'] = 'Short Description'
worksheet['E1'] = 'Status'
worksheet['F1'] = 'Weight'

# Generate 100 random incidents
for i in range(1, 101):
    # Set the incident number as a series
    worksheet.cell(row=i+1, column=1, value=i)
    
    # Set the priority to a random value between 1 and 5
    priority = random.choice(priorities)
    worksheet.cell(row=i+1, column=2, value=priority)
    
    # Set the date/time to a random value within the last 30 days
    now = datetime.now()
    delta = timedelta(days=30)
    random_date = now - random.uniform(0, 1) * delta
    formatted_date = random_date.strftime('%d/%m/%Y %H:%M:%S')
    worksheet.cell(row=i+1, column=3, value=formatted_date)
    
    # Set the short description to a random string of 10 characters
    description = ''.join(random.choices('abcdefghijklmnopqrstuvwxyz', k=10))
    worksheet.cell(row=i+1, column=4, value=description)
    
    # Set the status to New or In-progress before being Closed
    if i == 100:
        worksheet.cell(row=i+1, column=5, value='Closed')
    else:
        if random.choice([True, False]):
            worksheet.cell(row=i+1, column=5, value='New')
        else:
            worksheet.cell(row=i+1, column=5, value='In-progress')
    
    # Set the weight to a random value between 1 and 5
    weight = random.choice(weights)
    worksheet.cell(row=i+1, column=6, value=weight)

    # If the incident has a Resolved status, add a new entry with the same incident number
    if worksheet.cell(row=i+1, column=5).value == 'Resolved':
        j = i + 1
        while worksheet.cell(row=j, column=1).value == i:
            j += 1
        worksheet.insert_rows(j)
        for k in range(1, 7):
            worksheet.cell(row=j, column=k).value = worksheet.cell(row=i+1, column=k).value

# Save the workbook to a file
workbook.save('incidents3.xlsx')
